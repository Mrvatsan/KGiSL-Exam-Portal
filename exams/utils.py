import pandas as pd
import os
import io

def clean_and_parse_excel(file_obj):
    """
    Parses an uploaded Excel file (xlsx, xls, HTML content in xls).
    Returns a pandas DataFrame with standardized columns.
    """
    import logging
    import xlrd
    from io import BytesIO
    
    logging.basicConfig(filename='d:/adminstudent/debug.log', level=logging.DEBUG, 
                       format='%(asctime)s - %(message)s')
    
    df = None
    excel_error = None
    html_error = None
    
    # Get file name to determine type
    file_name = getattr(file_obj, 'name', '')
    logging.debug(f"Processing file: {file_name}")
    print(f"DEBUG: Processing file: {file_name}")
    
    # Read first few bytes to check content type
    file_obj.seek(0)
    first_bytes = file_obj.read(500)
    file_obj.seek(0)
    
    logging.debug(f"First 500 bytes: {first_bytes}")
    print(f"DEBUG: First 200 chars of file: {first_bytes[:200]}")
    
    is_html = b'<html' in first_bytes.lower() or b'<table' in first_bytes.lower() or b'<?xml' in first_bytes.lower()
    
    # Strategy 1: Try HTML parsing first if HTML detected
    if is_html:
        logging.debug("Detected HTML/XML content")
        print(f"DEBUG: Detected HTML/XML content in file")
        for engine in ['lxml', 'html5lib', 'bs4']:
            try:
                file_obj.seek(0)
                dfs = pd.read_html(file_obj, flavor=engine)
                if dfs and len(dfs) > 0:
                    df = dfs[0]
                    logging.debug(f"Successfully read HTML with {engine}: {len(df)} rows, {len(df.columns)} cols")
                    print(f"DEBUG: Successfully read HTML table with engine '{engine}': {len(df)} rows, {len(df.columns)} columns")
                    break
            except Exception as e:
                logging.debug(f"HTML read with {engine} failed: {str(e)}")
                print(f"DEBUG: HTML read with {engine} failed: {str(e)}")
                html_error = str(e)
    
    # Strategy 2: Try standard Excel parsing if not HTML or if HTML failed
    if df is None:
        logging.debug("Trying Excel parsing")
        try:
            file_obj.seek(0)
            if file_name.endswith('.xlsx'):
                df = pd.read_excel(file_obj, engine='openpyxl')
            elif file_name.endswith('.xls'):
                # For .xls files, try direct xlrd parsing with forgiving options
                try:
                    file_obj.seek(0)
                    file_contents = file_obj.read()
                    workbook = xlrd.open_workbook(file_contents=file_contents, 
                                                 formatting_info=False,
                                                 on_demand=True,
                                                 ignore_workbook_corruption=True)
                    sheet = workbook.sheet_by_index(0)
                    
                    # Convert to list of lists
                    data = []
                    for row_idx in range(sheet.nrows):
                        row = []
                        for col_idx in range(sheet.ncols):
                            cell = sheet.cell(row_idx, col_idx)
                            row.append(cell.value)
                        data.append(row)
                    
                    # Create DataFrame
                    if len(data) > 0:
                        df = pd.DataFrame(data[1:], columns=data[0])
                        logging.debug(f"Successfully read XLS with direct xlrd: {len(df)} rows, {len(df.columns)} cols")
                        print(f"DEBUG: Successfully read XLS with direct xlrd: {len(df)} rows, {len(df.columns)} columns")
                except Exception as xlrd_err:
                    logging.debug(f"Direct xlrd failed: {str(xlrd_err)}")
                    print(f"DEBUG: Direct xlrd failed: {str(xlrd_err)}")
                    raise xlrd_err
            else:
                # Try openpyxl first, then xlrd
                try:
                    df = pd.read_excel(file_obj, engine='openpyxl')
                except:
                    file_obj.seek(0)
                    df = pd.read_excel(file_obj, engine='xlrd')
            
            if df is not None:
                logging.debug(f"Successfully read Excel: {len(df)} rows, {len(df.columns)} cols")
                print(f"DEBUG: Successfully read Excel file: {len(df)} rows, {len(df.columns)} columns")
        except Exception as e:
            excel_error = str(e)
            logging.debug(f"Excel read failed: {excel_error}")
            print(f"DEBUG: Excel read failed: {excel_error}")

    if df is None:
        error_msg = f"Could not parse file. Ensure it is a valid Excel or HTML table file. Excel error: {excel_error}, HTML error: {html_error}"
        logging.debug(error_msg)
        print(f"DEBUG: {error_msg}")
        raise ValueError(error_msg)

    # Standardize Column Names - convert to lowercase for matching
    df.columns = df.columns.astype(str).str.lower().str.strip().str.replace(r'\s+', '', regex=True)
    
    print(f"DEBUG: Detected Columns in file: {list(df.columns)}")

    # Robust Fuzzy Mapping
    # We iterate over columns and try to match them to expected fields
    
    normalization_map = {}
    
    for col in df.columns:
        # Register No Matching - handle 'registerno', 'register_no', 'regno', etc.
        if 'register' in col or col == 'regno' or col == 'registerno':
            normalization_map[col] = 'register_no'
        elif 'rollno' in col or col == 'roll':
            normalization_map[col] = 'register_no'
            
        # Name Matching - handle 'studentname', 'name', 'student_name'
        elif ('student' in col and 'name' in col) or (col == 'name' or col == 'studentname'):
            normalization_map[col] = 'name'
            
        # Course Code - handle 'coursecode', 'course_code', 'subjectcode'
        elif 'coursecode' in col or 'subjectcode' in col or 'subcode' in col:
             normalization_map[col] = 'course_code'
             
        # Course Title - handle 'coursetitle', 'course_title', 'subjecttitle'
        elif 'coursetitle' in col or 'subjecttitle' in col or 'subjecttname' in col:
             normalization_map[col] = 'course_title'
             
        # Hall - handle 'examhallnumber', 'hallno', 'hall_no', 'room'
        elif 'hall' in col or 'room' in col:
            normalization_map[col] = 'hall_no'
            
        # Seat - handle 'examseatnumber', 'seatno', 'seat_no'
        elif 'seat' in col:
            normalization_map[col] = 'seat_no'
            
        # Session - handle 'examsession', 'session'
        elif 'session' in col:
             normalization_map[col] = 'session'
             
        # Date - handle 'examdate', 'exam_date', 'date'
        elif 'examdate' in col or (col == 'date'):
            normalization_map[col] = 'exam_date'

    # Apply mapping
    df = df.rename(columns=normalization_map)
    print(f"DEBUG: Mapped Columns: {list(df.columns)}")

    # Check for critical column
    if 'register_no' not in df.columns:
        # Fallback: if 'register_no' is missing but we have column 0, maybe use that? 
        # No, that's too risky.
        raise ValueError(f"Could not find 'Register No' column. Found columns: {list(df.columns)}")
        
    df = df.dropna(subset=['register_no'])
         
    return df


# ============ HALL TICKET UTILITY FUNCTIONS ============

def normalize_department_name(dept_name):
    """
    Normalize department name to canonical key.
    Handles variations in casing, spacing, and symbols.
    
    This ensures consistency between:
    - Register number department codes
    - Excel worksheet names
    - Database storage
    - Query matching
    
    Canonical Keys:
    AI&ML, AI&DS, CSE, ECE, IT, MECH, R&A, CSBS, CYS
    
    Examples:
    'AIML' → 'AI&ML'
    'AI ML' → 'AI&ML'
    'AI-ML' → 'AI&ML'
    'AI_ML' → 'AI&ML'
    'ai&ml' → 'AI&ML'
    'CSE' → 'CSE'
    """
    if not dept_name:
        return None
    
    # Convert to uppercase and remove extra whitespace
    normalized = str(dept_name).upper().strip()
    
    # Remove common separators to create a base form
    base_form = normalized.replace('&', '').replace('-', '').replace('_', '').replace(' ', '')
    
    # Map base forms to canonical department names
    canonical_map = {
        'AIML': 'AI&ML',
        'AIDS': 'AI&DS',
        'CSE': 'CSE',
        'ECE': 'ECE',
        'IT': 'IT',
        'MECH': 'MECH',
        'RA': 'R&A',
        'CSBS': 'CSBS',
        'CYS': 'CYS',
    }
    
    # Check if base form matches
    if base_form in canonical_map:
        return canonical_map[base_form]
    
    # If already in canonical form, return as-is
    if normalized in canonical_map.values():
        return normalized
    
    # No match found
    return None


def get_department_from_register_no(register_no):
    """
    Extract department from register number by searching for 3-letter department code.
    Returns department name or None if invalid.
    
    The department code can appear anywhere in the register number.
    This logic is year-independent and works with any register number format.
    
    Mapping:
    UAM → AI&ML
    UAD → AI&DS
    UCS → CSE
    UEC → ECE
    UME → MECH
    UIT → IT
    URA → R&A
    UCB → CSBS
    UCY → CYS
    
    Examples:
    711725UAM132 → AI&ML
    711724UAD217 → AI&DS
    711623UCS089 → CSE
    711726UEC210 → ECE
    """
    if not register_no:
        return None
    
    # Convert to uppercase for case-insensitive matching
    register_no_upper = str(register_no).upper()
    
    # Department code mapping (3-letter codes)
    dept_code_map = {
        'UAM': 'AI&ML',
        'UAD': 'AI&DS',
        'UCS': 'CSE',
        'UEC': 'ECE',
        'UME': 'MECH',
        'UIT': 'IT',
        'URA': 'R&A',
        'UCB': 'CSBS',
        'UCY': 'CYS',
    }
    
    # Search for any valid department code in the register number
    for code, department in dept_code_map.items():
        if code in register_no_upper:
            return department
    
    # No valid department code found
    return None


def parse_hall_ticket_excel(file_obj):
    """
    Parse hall ticket Excel with multiple department worksheets.
    Returns list of exam records ready for database insertion.
    
    Uses department normalization to handle naming variations in worksheets.
    All department names are normalized to canonical keys before storage.
    
    Expected structure:
    - Multiple worksheets (one per department)
    - Columns: Semester, CourseCode, CourseTitle, ExamDate, ExamSession
    """
    import openpyxl
    from datetime import datetime
    
    records = []
    
    # Load workbook
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj)
    
    # Process each worksheet
    for sheet_name in wb.sheetnames:
        # Normalize worksheet name to canonical department key
        department = normalize_department_name(sheet_name)
        
        if not department:
            print(f"WARNING: Unknown worksheet '{sheet_name}', skipping...")
            continue
        
        print(f"Processing worksheet '{sheet_name}' → Department: {department}")
        
        sheet = wb[sheet_name]
        
        # Read header row (assumed to be row 1)
        headers = [cell.value for cell in sheet[1]]
        
        # Normalize headers
        header_map = {}
        for idx, header in enumerate(headers):
            if header:
                normalized = str(header).lower().strip()
                if 'semester' in normalized:
                    header_map['semester'] = idx
                elif 'coursecode' in normalized or 'code' in normalized:
                    header_map['course_code'] = idx
                elif 'coursetitle' in normalized or 'title' in normalized:
                    header_map['course_title'] = idx
                elif 'examdate' in normalized or 'date' in normalized:
                    header_map['exam_date'] = idx
                elif 'session' in normalized:
                    header_map['session'] = idx
        
        # Validate required columns
        required = ['semester', 'course_code', 'course_title', 'exam_date', 'session']
        missing = [col for col in required if col not in header_map]
        if missing:
            print(f"WARNING: Worksheet '{sheet_name}' missing columns: {missing}, skipping...")
            continue
        
        # Read data rows (starting from row 2)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                semester = row[header_map['semester']]
                course_code = row[header_map['course_code']]
                course_title = row[header_map['course_title']]
                exam_date = row[header_map['exam_date']]
                session = row[header_map['session']]
                
                # Skip empty rows or rows with missing critical data
                if not semester or not course_code or not exam_date:
                    continue
                
                # Parse exam date
                if isinstance(exam_date, datetime):
                    exam_date = exam_date.date()
                elif isinstance(exam_date, str):
                    exam_date = pd.to_datetime(exam_date).date()
                else:
                    # Skip rows with invalid date format
                    continue
                
                # Clean session value
                session = str(session).strip().upper()
                
                # Store with canonical department name
                records.append({
                    'department': department,  # Already normalized
                    'semester': str(semester).strip(),
                    'course_code': str(course_code).strip(),
                    'course_title': str(course_title).strip(),
                    'exam_date': exam_date,
                    'session': session,
                })
            except Exception as e:
                print(f"WARNING: Error parsing row in '{sheet_name}': {e}")
                continue
    
    return records
